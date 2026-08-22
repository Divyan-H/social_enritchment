"""
enrich_handles.py — reads an Excel file where a column (e.g. "Social Handle"
or "Post Link") contains hyperlinks, URLs, or handles for Instagram and/or
YouTube, extracts the underlying hyperlink targets, filters out plain text/noise,
and enriches the data using fast batched scraping (mirroring enrich.py).

Features:
- Extracts actual hyperlink target URLs from Excel cells (direct cell hyperlinks,
  sheet-level hyperlinks, and =HYPERLINK() formulas).
- Intelligently recognizes Instagram post/reel links, Instagram profile links,
  YouTube video/shorts links, YouTube channel links, and handles.
- Skips non-link text, headers (e.g. '2022-23'), category names, blank cells,
  and whitespace without sending invalid handles to Apify.
- Fast batched Apify scraping for Instagram posts & profiles and YouTube channels.
- Fast, non-blocking YouTube scraper (extract_flat + socket_timeout + hard timeouts).
- Offline keyword-based genre classification.
- Robust data extraction handling diverse response shapes (including nested URLs/dicts).
- Full output matching enrich.py enr_* schema.

Usage:
    python enrich_handles.py input.xlsx
    python enrich_handles.py input.xlsx --out result.xlsx
    python enrich_handles.py input.xlsx --handles-col "Social Handle"
    python enrich_handles.py input.xlsx --max 20
    python enrich_handles.py input.xlsx --rows 2,5,10-15
"""
import argparse
import os
import re
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Optional, Tuple, Dict, Any
from urllib.parse import urlparse, parse_qs

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import openpyxl.utils

import apify_client
import scraper
from genre import classify_genre
from enrich import parse_rows_arg, OUTPUT_FIELDS, STATUS_COL, ALL_ENR_COLS


# ── Hyperlink-aware Excel reader ────────────────────────────────────────────

def _extract_col_with_hyperlinks(filepath: str, sheet, col_name: str, nrows: int) -> list:
    """Extract cell values and hyperlink targets from Excel worksheet.
    Returns a list of length `nrows` where each element is the hyperlink target URL
    if present, or the raw cell value.
    """
    try:
        wb = load_workbook(filepath, data_only=False)
    except Exception as e:
        print(f"Warning: could not open workbook with openpyxl ({e}); falling back to standard values.")
        return [None] * nrows

    try:
        if isinstance(sheet, int):
            ws = wb.worksheets[sheet]
        elif isinstance(sheet, str) and sheet in wb.sheetnames:
            ws = wb[sheet]
        elif isinstance(sheet, str) and sheet.isdigit() and int(sheet) < len(wb.worksheets):
            ws = wb.worksheets[int(sheet)]
        else:
            ws = wb.active
    except Exception:
        wb.close()
        return [None] * nrows

    # Locate target column by header name (row 1, case-insensitive)
    col_idx = None
    header_row = next(ws.iter_rows(min_row=1, max_row=1), [])
    for cell in header_row:
        if str(cell.value or "").strip().lower() == str(col_name).strip().lower():
            col_idx = cell.column
            break

    if col_idx is None:
        wb.close()
        return [None] * nrows

    # Build row_idx -> target map from sheet-level hyperlinks
    hl_map = {}
    if hasattr(ws, "hyperlinks"):
        for hl in ws.hyperlinks:
            if hasattr(hl, "ref") and hasattr(hl, "target") and hl.target:
                ref = hl.ref
                if ":" in ref:
                    try:
                        min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(ref)
                        if min_col <= col_idx <= max_col:
                            for r in range(min_row, max_row + 1):
                                hl_map[r] = hl.target
                    except Exception:
                        pass
                else:
                    try:
                        c_idx, r_idx = openpyxl.utils.coordinate_to_tuple(ref)
                        if c_idx == col_idx:
                            hl_map[r_idx] = hl.target
                    except Exception:
                        pass

    _hl_formula_re = re.compile(r'HYPERLINK\(\s*["\']([^"\']+)["\']', re.I)

    values = []
    for row_idx in range(2, nrows + 2):  # row 1 is header
        cell = ws.cell(row=row_idx, column=col_idx)
        val = None

        # 1. Direct cell hyperlink
        if cell.hyperlink and getattr(cell.hyperlink, "target", None):
            val = cell.hyperlink.target

        # 2. Sheet-level hyperlink
        if not val and row_idx in hl_map:
            val = hl_map[row_idx]

        # 3. Formula =HYPERLINK("url", ...)
        if not val and isinstance(cell.value, str):
            m = _hl_formula_re.search(cell.value)
            if m:
                val = m.group(1)

        # 4. Fallback to cell value
        if not val:
            val = cell.value

        values.append(val)

    wb.close()
    return values


# ── URL & Handle Parsing & Classification ───────────────────────────────────

_URL_RE = re.compile(r"https?://[^\s,;\"'<>]+", re.I)
_IG_POST_PATH_RE = re.compile(r"^/(?:p|reel|reels|tv)/([A-Za-z0-9_-]+)", re.I)

_IG_LABEL_RE = re.compile(r"^\s*(?:insta(?:gram)?|ig)\s*[:\-]?\s*", re.I)
_YT_LABEL_RE = re.compile(r"^\s*(?:you\s*tube|yt)\s*[:\-]?\s*", re.I)
_DELIM_RE = re.compile(r"[,;|]+|\s+and\s+|\s+/\s+|\n+", re.I)

_BLANK_VALUES = {"", "nan", "none", "-", "n/a", "na", "null", "nil", "unknown", "skip", "undefined"}
_YEAR_OR_RANGE_RE = re.compile(r"^\d{2,4}(?:[\s\-_/]+\d{2,4})+$")


def _clean_handle(h: str) -> str:
    if not h:
        return ""
    h = str(h).strip().lstrip("@")
    h = h.split("?")[0].split("#")[0].strip().rstrip("/")
    return h


def _clean_url(url: str) -> str:
    """Normalize and clean URL by stripping tracking parameters."""
    if not url:
        return ""
    u = str(url).strip()
    # Strip tracking params from instagram/youtube URLs
    if "instagram.com" in u or "instagr.am" in u:
        u = u.split("?")[0].rstrip("/") + "/"
    elif "youtube.com" in u or "youtu.be" in u:
        if "youtu.be/" in u:
            vid_id = u.split("youtu.be/")[-1].split("?")[0].split("/")[0]
            u = f"https://www.youtube.com/watch?v={vid_id}"
        elif "watch" in u:
            parsed = urlparse(u)
            params = parse_qs(parsed.query)
            vid = params.get("v", [""])[0]
            if vid:
                u = f"https://www.youtube.com/watch?v={vid}"
        else:
            u = u.split("?")[0].rstrip("/")
    return u


def _is_valid_ig_handle(h: str) -> bool:
    """Check if string is a plausible Instagram handle (letters, numbers, periods, underscores, 1-30 chars)."""
    if not h or len(h) > 30:
        return False
    if _YEAR_OR_RANGE_RE.match(h):
        return False
    if " " in h or "-" in h or "/" in h or "\\" in h:
        return False
    return bool(re.match(r"^[A-Za-z0-9_.]{1,30}$", h)) and not h.isdigit()


def _is_valid_yt_handle(h: str) -> bool:
    """Check if string is a plausible YouTube handle."""
    if not h or len(h) > 50:
        return False
    if _YEAR_OR_RANGE_RE.match(h):
        return False
    if " " in h:
        return False
    clean = h.lstrip("@")
    return bool(re.match(r"^[A-Za-z0-9_.\-]{1,50}$", clean)) and not clean.isdigit()


def parse_target_item(raw_val: Any, default_platform: str = "instagram") -> dict:
    """Classify a cell value into structured target(s).
    Returns dict:
      {
        "ig_post_url": str or None,
        "ig_profile_handle": str or None,
        "yt_video_url": str or None,
        "yt_channel_handle": str or None,
        "is_skipped": bool,
        "skip_reason": str or None
      }
    """
    res = {
        "ig_post_url": None,
        "ig_profile_handle": None,
        "yt_video_url": None,
        "yt_channel_handle": None,
        "is_skipped": False,
        "skip_reason": None,
    }

    if raw_val is None:
        res["is_skipped"] = True
        res["skip_reason"] = "no link"
        return res

    text = str(raw_val).strip()
    if not text or text.lower() in _BLANK_VALUES:
        res["is_skipped"] = True
        res["skip_reason"] = "no link"
        return res

    if _YEAR_OR_RANGE_RE.match(text):
        res["is_skipped"] = True
        res["skip_reason"] = "skipped (header/text)"
        return res

    # 1. Extract URLs
    urls_found = _URL_RE.findall(text)
    for u in urls_found:
        u_clean = _clean_url(u)
        host = urlparse(u).netloc.lower().lstrip("www.")
        path = urlparse(u).path

        if "instagram" in host or "instagr.am" in host:
            post_m = _IG_POST_PATH_RE.search(path)
            if post_m:
                if not res["ig_post_url"]:
                    res["ig_post_url"] = u_clean
            else:
                # Profile URL
                seg = path.strip("/").split("/")[0] if path.strip("/") else ""
                if seg and seg.lower() not in ("p", "reel", "reels", "tv", "stories", "explore", "direct", "accounts"):
                    h = _clean_handle(seg)
                    if _is_valid_ig_handle(h) and not res["ig_profile_handle"]:
                        res["ig_profile_handle"] = h
        elif "youtube" in host or "youtu.be" in host:
            if "youtu.be" in host or "/watch" in path or "/shorts" in path or "/embed" in path:
                if not res["yt_video_url"]:
                    res["yt_video_url"] = u_clean
            else:
                # Channel URL
                parts = path.strip("/").split("/")
                if parts and parts[0].startswith("@"):
                    res["yt_channel_handle"] = parts[0]
                elif len(parts) > 1 and parts[0].lower() in ("c", "user", "channel"):
                    res["yt_channel_handle"] = parts[1]
                elif parts and parts[0] and parts[0].lower() not in ("watch", "shorts", "embed", "live"):
                    res["yt_channel_handle"] = parts[0]

        text = text.replace(u, " ")

    # If URLs satisfied either IG or YT, return
    if any((res["ig_post_url"], res["ig_profile_handle"], res["yt_video_url"], res["yt_channel_handle"])):
        return res

    # 2. Check for labeled handles or plain tokens
    tokens = [t.strip() for t in _DELIM_RE.split(text) if t.strip()]
    unlabeled = []

    for tok in tokens:
        if tok.lower() in _BLANK_VALUES or _YEAR_OR_RANGE_RE.match(tok):
            continue

        ig_m = _IG_LABEL_RE.match(tok)
        if ig_m:
            h = _clean_handle(tok[ig_m.end():])
            if _is_valid_ig_handle(h):
                res["ig_profile_handle"] = h
            continue

        yt_m = _YT_LABEL_RE.match(tok)
        if yt_m:
            h = _clean_handle(tok[yt_m.end():])
            if _is_valid_yt_handle(h):
                res["yt_channel_handle"] = h
            continue

        if tok.startswith("@"):
            clean_tok = _clean_handle(tok)
            if _is_valid_ig_handle(clean_tok):
                unlabeled.append(clean_tok)
        elif _is_valid_ig_handle(tok):
            unlabeled.append(tok)

    if not any((res["ig_post_url"], res["ig_profile_handle"], res["yt_video_url"], res["yt_channel_handle"])):
        if len(unlabeled) == 1:
            if default_platform == "instagram" and _is_valid_ig_handle(unlabeled[0]):
                res["ig_profile_handle"] = unlabeled[0]
            elif default_platform == "youtube" and _is_valid_yt_handle(unlabeled[0]):
                res["yt_channel_handle"] = unlabeled[0]
            else:
                res["is_skipped"] = True
                res["skip_reason"] = "skipped (no valid link/handle)"
        elif len(unlabeled) >= 2:
            if _is_valid_ig_handle(unlabeled[0]):
                res["ig_profile_handle"] = unlabeled[0]
            if _is_valid_yt_handle(unlabeled[1]):
                res["yt_channel_handle"] = unlabeled[1]
        else:
            res["is_skipped"] = True
            res["skip_reason"] = "skipped (no valid link/handle)"

    return res


# ── Scraping Workers & Batch Logic ──────────────────────────────────────────

def _safe_external_url(ext: Any) -> str:
    """Safely convert external_url field (which can be a string, dict, list of strings/dicts) to a clean string."""
    if not ext:
        return ""
    if isinstance(ext, str):
        return ext
    if isinstance(ext, dict):
        return ext.get("url") or ext.get("lynx_url") or str(ext)
    if isinstance(ext, list):
        items = []
        for item in ext:
            if isinstance(item, str):
                items.append(item)
            elif isinstance(item, dict):
                u = item.get("url") or item.get("lynx_url") or str(item)
                if u:
                    items.append(u)
            elif item is not None:
                items.append(str(item))
        return ", ".join(items) if items else ""
    return str(ext)


def _apply_profile_fields(sp: scraper.ScrapedPost, prof: dict):
    """Safely apply profile fields from Apify without crashing on nested types."""
    sp.followers = prof.get("followers")
    sp.following = prof.get("following")
    sp.posts_count = prof.get("posts_count")
    sp.biography = prof.get("biography") or ""
    sp.full_name = prof.get("full_name") or ""
    sp.is_verified = prof.get("is_verified")
    sp.is_private = prof.get("is_private")
    sp.is_business_account = prof.get("is_business_account")
    sp.business_category = prof.get("business_category") or ""
    sp.external_url = _safe_external_url(prof.get("external_url"))


def _row_values_from_result(result) -> dict:
    out = {}
    for attr, col in OUTPUT_FIELDS:
        val = getattr(result, attr, None) if result else None
        if isinstance(val, list):
            val = ", ".join(str(v) for v in val) if val else None
        elif val == "":
            val = None
        out[col] = val
    if result:
        out[STATUS_COL] = "ok" if getattr(result, "ok", False) else f"failed: {getattr(result, 'error', 'unknown error')}"
    else:
        out[STATUS_COL] = "failed: no result"
    return out


def _empty_row(status: str) -> dict:
    out = {col: None for _, col in OUTPUT_FIELDS}
    out[STATUS_COL] = status
    return out


def _scrape_instagram_posts_batch(urls: list, apify_token: str) -> dict:
    """Scrape Instagram post URLs in batch + owning profile data safely."""
    urls = list(dict.fromkeys(urls))
    results = {u: scraper.ScrapedPost(url=u, platform="instagram") for u in urls}

    posts_by_url = apify_client.fetch_posts_batch(urls, apify_token)

    handles_needed = set()
    for u in urls:
        raw = posts_by_url.get(u)
        post_fields = apify_client.extract_post_fields(raw) if raw else {}
        sp = results[u]
        if post_fields.get("description") or post_fields.get("handle"):
            sp.ok = True
            sp.source = "apify"
            scraper._apply_apify_post_fields(sp, post_fields)
            if sp.handle:
                handles_needed.add(sp.handle)

    if handles_needed:
        profiles_by_handle = apify_client.fetch_profiles_batch(list(handles_needed), apify_token)
        for u in urls:
            sp = results[u]
            if sp.handle and sp.handle in profiles_by_handle:
                prof_raw = profiles_by_handle[sp.handle]
                if prof_raw:
                    _apply_profile_fields(sp, apify_client.extract_profile_fields(prof_raw))

    # Fallback only for the URLs Apify gave nothing for
    missed = [u for u in urls if not results[u].ok]
    for u in missed:
        sp = results[u]
        sp = scraper._scrape_instagram_free_fallback(sp, u, ["apify: no item returned for this URL in the batch"])
        results[u] = sp

    for u in urls:
        sp = results[u]
        if sp.description or sp.handle:
            genre, conf = classify_genre(sp.description, sp.handle)
            sp.genre = genre
            sp.genre_confidence = conf

    return results


def _scrape_instagram_profiles_batch(handles: list, apify_token: Optional[str]) -> dict:
    """Fetch Instagram profile data for list of handles -> {handle: ScrapedPost}."""
    handles = list(dict.fromkeys(handles))
    results = {}
    if not handles:
        return results

    misses = list(handles)
    if apify_token:
        print(f"Batch-scraping {len(handles)} unique Instagram profile(s) via Apify...")
        raw_by_handle = apify_client.fetch_profiles_batch(handles, apify_token)
        misses = []
        for h in handles:
            raw = raw_by_handle.get(h)
            if raw:
                fields = apify_client.extract_profile_fields(raw)
                sp = scraper.ScrapedPost(url=f"https://www.instagram.com/{h}/", platform="instagram", ok=True, source="apify")
                sp.handle = h
                _apply_profile_fields(sp, fields)
                genre, conf = classify_genre(sp.biography, sp.handle)
                sp.genre = genre
                sp.genre_confidence = conf
                results[h] = sp
            else:
                misses.append(h)

        if misses:
            print(f"Apify returned nothing for {len(misses)} handle(s) - trying free fallback...")

    for h in misses:
        stats = scraper._fetch_instagram_profile_stats(h)
        sp = scraper.ScrapedPost(url=f"https://www.instagram.com/{h}/", platform="instagram")
        sp.handle = h
        if stats and (stats.get("followers") is not None or stats.get("following") is not None):
            sp.ok = True
            sp.source = "free_fallback"
            sp.followers = stats.get("followers")
            sp.following = stats.get("following")
        else:
            reason = scraper.get_profile_fail_reason(h) or "no data returned"
            sp.error = reason
        results[h] = sp

    return results


# ── YouTube Fast & Non-Blocking Fetchers ────────────────────────────────────

def _safe_fetch_youtube_channel_info(handle_or_url: str) -> Optional[dict]:
    """Extract YouTube channel info fast without hanging or parsing all videos."""
    if not handle_or_url:
        return None
    h = str(handle_or_url).strip()
    if h.startswith("http"):
        url = h
    elif h.startswith("@"):
        url = f"https://www.youtube.com/{h}"
    else:
        url = f"https://www.youtube.com/@{h}"

    # 1. Try yt-dlp with extract_flat=True and 8s socket timeout
    try:
        import yt_dlp
        opts = {
            "quiet": True,
            "no_warnings": True,
            "skip_download": True,
            "ignoreerrors": True,
            "extract_flat": True,          # Prevents downloading entire playlist/channel videos
            "playlist_items": "0",         # Fetches channel info only
            "socket_timeout": 8,
            "logger": type("NullLogger", (), {
                "debug": lambda s, m: None, "info": lambda s, m: None,
                "warning": lambda s, m: None, "error": lambda s, m: None,
            })(),
        }
        with yt_dlp.YoutubeDL(opts) as ydl:
            info = ydl.extract_info(url, download=False)
        if info and (info.get("channel") or info.get("uploader") or info.get("channel_follower_count")):
            return {
                "channel_name": info.get("channel") or info.get("uploader") or info.get("uploader_id") or "",
                "subscribers": info.get("channel_follower_count"),
                "description": (info.get("description") or "")[:1500],
                "video_count": None,
            }
    except Exception:
        pass

    # 2. Fast HTML meta tag fallback
    try:
        r = requests.get(url, headers=scraper.HEADERS, timeout=6)
        if r.status_code == 200:
            soup = BeautifulSoup(r.text, "html.parser")
            og_title = ""
            og_desc = ""
            for m in soup.find_all("meta"):
                name = (m.get("name") or m.get("property") or "").lower()
                if name in ("og:title", "title"):
                    og_title = m.get("content") or ""
                elif name in ("og:description", "description"):
                    og_desc = m.get("content") or ""
            if og_title or og_desc:
                return {
                    "channel_name": og_title,
                    "subscribers": None,
                    "description": og_desc[:1500],
                    "video_count": None,
                }
    except Exception:
        pass

    return None


def _safe_fetch_youtube_video_info(url: str) -> scraper.ScrapedPost:
    """Scrape a YouTube video URL with fast timeout and error protection."""
    sp = scraper.ScrapedPost(url=url, platform="youtube")
    try:
        import yt_dlp
        opts = {
            "quiet": True,
            "no_warnings": True,
            "skip_download": True,
            "ignoreerrors": True,
            "socket_timeout": 8,
            "logger": type("NullLogger", (), {
                "debug": lambda s, m: None, "info": lambda s, m: None,
                "warning": lambda s, m: None, "error": lambda s, m: None,
            })(),
        }
        with yt_dlp.YoutubeDL(opts) as ydl:
            info = ydl.extract_info(url, download=False)
        if info:
            sp.ok = True
            sp.source = "ytdlp"
            sp.description = (info.get("description") or "")[:1500]
            sp.handle = info.get("uploader_id") or info.get("uploader") or info.get("channel") or ""
            sp.subscribers = info.get("channel_follower_count")
            genre, conf = classify_genre(sp.description, sp.handle)
            sp.genre = genre
            sp.genre_confidence = conf
        else:
            sp.error = "yt-dlp returned no info"
    except Exception as e:
        sp.error = f"{type(e).__name__}: {e}"
    return sp


def _scrape_youtube_channels_batch(handles: list, workers: int, apify_token: Optional[str]) -> dict:
    """Fetch YouTube channel data for list of handles/URLs -> {handle: ScrapedPost}."""
    handles = list(dict.fromkeys(handles))
    results = {}
    if not handles:
        return results

    misses = list(handles)
    if apify_token:
        print(f"Batch-scraping {len(handles)} unique YouTube channel(s) via Apify...")
        try:
            raw_by_handle = apify_client.fetch_youtube_channels_batch(handles, apify_token)
        except Exception as e:
            print(f"Warning: Apify YouTube channel scrape error ({e}) - falling back to yt-dlp")
            raw_by_handle = {}

        misses = []
        for h in handles:
            raw = raw_by_handle.get(h)
            if raw:
                fields = apify_client.extract_youtube_channel_fields(raw)
                sp = scraper.ScrapedPost(url=f"https://www.youtube.com/{h}", platform="youtube", ok=True, source="apify")
                sp.handle = fields.get("handle") or h
                sp.full_name = fields.get("channel_name") or ""
                sp.subscribers = fields.get("subscribers")
                sp.posts_count = fields.get("video_count")
                sp.description = fields.get("description") or ""
                genre, conf = classify_genre(sp.description, sp.full_name or sp.handle)
                sp.genre = genre
                sp.genre_confidence = conf
                results[h] = sp
            else:
                misses.append(h)

        if misses:
            print(f"Apify returned nothing for {len(misses)} YouTube channel(s) — falling back to fast yt-dlp...")

    if misses:
        print(f"Scraping {len(misses)} YouTube channel(s) via fast yt-dlp with {workers} parallel workers...")
        with ThreadPoolExecutor(max_workers=workers) as ex:
            future_to_h = {ex.submit(_safe_fetch_youtube_channel_info, h): h for h in misses}
            for fut in as_completed(future_to_h):
                h = future_to_h[fut]
                sp = scraper.ScrapedPost(url=f"https://www.youtube.com/{h}", platform="youtube")
                try:
                    info = fut.result(timeout=12)  # Strict 12-second timeout
                except Exception as e:
                    sp.error = f"timeout/error: {e}"
                    results[h] = sp
                    continue

                if info and (info.get("subscribers") is not None or info.get("channel_name")):
                    sp.ok = True
                    sp.source = "ytdlp"
                    sp.handle = h
                    sp.full_name = info.get("channel_name") or ""
                    sp.subscribers = info.get("subscribers")
                    sp.description = info.get("description") or ""
                    genre, conf = classify_genre(sp.description, sp.full_name or sp.handle)
                    sp.genre = genre
                    sp.genre_confidence = conf
                else:
                    sp.error = "no channel data returned (bad handle or channel private)"
                results[h] = sp

    return results


# ── Main Entrypoint ─────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(
        description="Enrich an Excel sheet containing Instagram/YouTube links, hyperlinks, or handles."
    )
    ap.add_argument("input", help="Path to input .xlsx file")
    ap.add_argument("--out", default=None, help="Output .xlsx path (default: <input>_enriched.xlsx)")
    ap.add_argument("--sheet", default=0, help="Sheet name or index to read (default: first sheet)")
    ap.add_argument("--handles-col", default=None, help="Column holding links or handles (default: auto-detect)")
    ap.add_argument("--link-col", default=None, help="Alias for --handles-col")
    ap.add_argument("--default-platform", choices=["instagram", "youtube", "skip"], default="instagram",
                    help="Platform for lone unlabeled handles without URL/prefix. Default: instagram.")
    ap.add_argument("--max", type=int, default=None, help="Cap number of rows to scrape (for testing)")
    ap.add_argument("--rows", default=None,
                    help="Only scrape these rows, 1-indexed with row 1 = header (e.g. '2,5,10-15').")
    ap.add_argument("--apify-token", default=None,
                    help="Apify API token for Instagram & YouTube. Prefer APIFY_API_TOKEN env var.")
    ap.add_argument("--youtube-workers", type=int, default=8,
                    help="How many YouTube items to scrape in parallel via yt-dlp (default 8).")
    args = ap.parse_args()

    apify_token = args.apify_token or os.environ.get("APIFY_API_TOKEN")
    if not apify_token:
        print("NOTE: no Apify token found (set APIFY_API_TOKEN or pass --apify-token). "
              "Falling back to free unauthenticated scrapes where available.")

    df = pd.read_excel(args.input, sheet_name=args.sheet)

    # Determine target column
    target_col = args.handles_col or args.link_col
    if not target_col:
        for candidate in ["Social Handle", "Post Link", "Link", "URL", "Handle", "Handles"]:
            if candidate in df.columns:
                target_col = candidate
                break
        if not target_col:
            target_col = df.columns[0]
            print(f"Note: no standard handle/link column name found, using first column '{target_col}'.")

    if target_col not in df.columns:
        print(f"ERROR: column '{target_col}' not found. Available columns: {list(df.columns)}")
        sys.exit(1)

    # Resolve hyperlinks from Excel sheet
    resolved_hyperlinks = _extract_col_with_hyperlinks(args.input, args.sheet, target_col, len(df))
    n_links = sum(1 for v in resolved_hyperlinks if v and str(v).startswith("http"))
    if n_links:
        print(f"Resolved {n_links} hyperlink URL(s) in column '{target_col}'.")

    row_selection = None
    if args.rows:
        try:
            row_selection = parse_rows_arg(args.rows, len(df))
        except ValueError as e:
            print(f"ERROR: --rows problem: {e}")
            sys.exit(1)
        print(f"--rows given: only scraping {len(row_selection)} row(s), leaving the rest untouched.")

    n = len(df) if args.max is None else min(args.max, len(df))

    # Preserve existing enr_* columns if already present
    existing = {col: (df[col].tolist() if col in df.columns else [None] * len(df)) for col in ALL_ENR_COLS}
    results_by_row = {}

    # ── Parse each row to categorize tasks ──────────────────────────────────
    parsed_rows = {}
    ig_posts_to_scrape = []       # (row_idx, post_url)
    ig_profiles_to_scrape = []    # (row_idx, handle)
    yt_videos_to_scrape = []      # (row_idx, video_url)
    yt_channels_to_scrape = []    # (row_idx, channel_handle)

    for i in range(len(df)):
        if row_selection is not None and i not in row_selection:
            prior = {col: existing[col][i] for col in ALL_ENR_COLS}
            if pd.isna(prior[STATUS_COL]):
                prior[STATUS_COL] = "skipped (not in --rows selection)"
            results_by_row[i] = prior
            continue

        if i >= n:
            results_by_row[i] = _empty_row("skipped (--max reached)")
            continue

        # Check resolved hyperlink first, then raw dataframe cell
        cell_val = resolved_hyperlinks[i] if i < len(resolved_hyperlinks) and resolved_hyperlinks[i] is not None else df.iloc[i].get(target_col, "")
        parsed = parse_target_item(cell_val, default_platform=args.default_platform)
        parsed_rows[i] = parsed

        if parsed["is_skipped"]:
            results_by_row[i] = _empty_row(parsed["skip_reason"] or "skipped")
            continue

        if parsed["ig_post_url"]:
            ig_posts_to_scrape.append((i, parsed["ig_post_url"]))
        elif parsed["ig_profile_handle"]:
            ig_profiles_to_scrape.append((i, parsed["ig_profile_handle"]))

        if parsed["yt_video_url"]:
            yt_videos_to_scrape.append((i, parsed["yt_video_url"]))
        elif parsed["yt_channel_handle"]:
            yt_channels_to_scrape.append((i, parsed["yt_channel_handle"]))

    total_active = len(ig_posts_to_scrape) + len(ig_profiles_to_scrape) + len(yt_videos_to_scrape) + len(yt_channels_to_scrape)
    print(f"Processing {len(parsed_rows)} row(s): {len(ig_posts_to_scrape)} Instagram post(s), "
          f"{len(ig_profiles_to_scrape)} Instagram profile(s), {len(yt_videos_to_scrape)} YouTube video(s), "
          f"{len(yt_channels_to_scrape)} YouTube channel(s).")

    # ── 1. Batch scrape Instagram Posts ─────────────────────────────────────
    ig_post_results = {}
    if ig_posts_to_scrape:
        ig_post_urls = [u for _, u in ig_posts_to_scrape]
        unique_urls = list(dict.fromkeys(ig_post_urls))
        if apify_token:
            print(f"Batch-scraping {len(unique_urls)} unique Instagram post URL(s) via Apify...")
            ig_post_results = _scrape_instagram_posts_batch(unique_urls, apify_token)
        else:
            ig_post_results = {u: scraper.scrape_post(u, "instagram") for u in unique_urls}

    # ── 2. Batch scrape Instagram Profiles ──────────────────────────────────
    ig_profile_results = {}
    if ig_profiles_to_scrape:
        ig_handles = [h for _, h in ig_profiles_to_scrape]
        unique_handles = list(dict.fromkeys(ig_handles))
        ig_profile_results = _scrape_instagram_profiles_batch(unique_handles, apify_token)

    # ── 3. Scrape YouTube Videos (Parallel) ──────────────────────────────────
    yt_video_results = {}
    if yt_videos_to_scrape:
        unique_vids = list(dict.fromkeys([u for _, u in yt_videos_to_scrape]))
        print(f"Scraping {len(unique_vids)} unique YouTube video(s) with {args.youtube_workers} parallel workers...")
        with ThreadPoolExecutor(max_workers=args.youtube_workers) as ex:
            fut_map = {ex.submit(_safe_fetch_youtube_video_info, u): u for u in unique_vids}
            for fut in as_completed(fut_map):
                u = fut_map[fut]
                try:
                    yt_video_results[u] = fut.result(timeout=15)
                except Exception as e:
                    sp = scraper.ScrapedPost(url=u, platform="youtube", error=f"timeout/error: {e}")
                    yt_video_results[u] = sp

    # ── 4. Scrape YouTube Channels ──────────────────────────────────────────
    yt_channel_results = {}
    if yt_channels_to_scrape:
        unique_channels = list(dict.fromkeys([h for _, h in yt_channels_to_scrape]))
        yt_channel_results = _scrape_youtube_channels_batch(unique_channels, args.youtube_workers, apify_token)

    # ── 5. Assemble row results ─────────────────────────────────────────────
    for i, parsed in parsed_rows.items():
        if i in results_by_row:
            continue  # Already filled (e.g. skipped)

        res_obj = None
        # Primary result from IG or YT
        if parsed.get("ig_post_url") and parsed["ig_post_url"] in ig_post_results:
            res_obj = ig_post_results[parsed["ig_post_url"]]
        elif parsed.get("ig_profile_handle") and parsed["ig_profile_handle"] in ig_profile_results:
            res_obj = ig_profile_results[parsed["ig_profile_handle"]]
        elif parsed.get("yt_video_url") and parsed["yt_video_url"] in yt_video_results:
            res_obj = yt_video_results[parsed["yt_video_url"]]
        elif parsed.get("yt_channel_handle") and parsed["yt_channel_handle"] in yt_channel_results:
            res_obj = yt_channel_results[parsed["yt_channel_handle"]]

        if res_obj:
            results_by_row[i] = _row_values_from_result(res_obj)
        else:
            results_by_row[i] = _empty_row("failed: no data returned")

    # Populate dataframe
    for col in ALL_ENR_COLS:
        df[col] = [results_by_row[i][col] for i in range(len(df))]

    out_path = args.out or args.input.rsplit(".", 1)[0] + "_enriched.xlsx"
    df.to_excel(out_path, index=False)
    n_processed = len(parsed_rows)
    print(f"\nDone. Wrote {out_path} ({n_processed} rows processed, {len(df) - n_processed} skipped/untouched)")


if __name__ == "__main__":
    main()